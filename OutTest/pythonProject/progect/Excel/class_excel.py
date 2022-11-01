# Excel 处理测试数据测试结果
# pip install openpyxl
# pip install ddt

#将数据存到Excel里面python去操作Excel
#1：只支持这种格式  .xlsx    openpyxl只支持这种格式
#2：老老实实收工创建

#1打开文件
# wb=load_workbook("test1.xlsx")
# sheet=wb["Sheet1"]
#3定位单元格 行列值
# res=sheet.cell(1,1).value
# print("最大行：{}".format(sheet.max_row))#求表单的最大行
# print("最大列：{}".format(sheet.max_column))#求表单的最大列

#
from openpyxl import load_workbook
from progect.Excel.case_1 import ReadConfig

#方法1
class DoExcel:
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name

    def get_data(self,mode="all"):

        mode = ReadConfig().read_config("case.config", "MODE", "mode")

        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]
        test_data=[]
        for i in range(1,sheet.max_row+1):
            sub_data = {}
            sub_data["url"]=sheet.cell(i,1).value
            sub_data["data"]=sheet.cell(i,2).value
            sub_data["expected"]=sheet.cell(i,3).value
            sub_data["method"]=sheet.cell(i,4).value
            test_data.append(sub_data)


        if mode=="all":
            final_data=test_data
        else:
            final_data=[]
            for item in test_data:#对test_data所有数据进行遍历
                if item["case_id"] in eval(mode):
                    final_data.append(item)

        return final_data#返回获取到的数据

#方法2
class DoExcel2:
    def __init__(self,file_name,sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.sheet_obj=load_workbook(self.file_name)[self.sheet_name]
        self.max_row=self.sheet_obj.max_row
        #获取一个表单的对象

    def get_data(self,i,j):#根据传入坐标进行取值
        return self.sheet_obj.cell(i,j).value

#方法3
class DoExcel3:
    def __init__(self,file_name,sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def get_header(self):
        wb=load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        header = []#存储标题
        for j in range(1,sheet.max_column+1):
            header.append(sheet.cell(1,j).value)
        return header

    def get_data(self):
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]

        header=self.get_header()#拿到headel
        test_data=[]

        for i in range(2,sheet.max_row+1):#第一行是标题  从第二行开始获取值
            sub_data = {}
            for j in range(1,sheet.max_column+1):#列
                sub_data[header[j-1]]=sheet.cell(i,j).value#下标注从0开始  所以减1
            test_data.append(sub_data)
        return test_data


if __name__ == '__main__':
    # print(DoExcel("test1.xlsx","Sheet1").get_data())#方法1拿到全部
    # print(DoExcel2("test1.xlsx", "Sheet1").get_data(1,2))#方法2拿到全部
    print(DoExcel3("test1.xlsx", "Sheet2").get_data())#方法3


