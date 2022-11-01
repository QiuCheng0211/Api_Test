from openpyxl import load_workbook
from tools.read_config import ReadConfig
from tools import project_path
from tools.get_data import GetData

""""
使用excel处理数据
"""
class DoExcel:

    @classmethod
    def get_data(cls,file_name):
        wb = load_workbook(file_name)  # 参数化传参 传入文件路径以和文件名
        mode = eval(ReadConfig.get_config(project_path.case_config_path,"MODE","mode")) #表单的页签

        # # 默认全部加载
        #         test_data=[] #列表
        #         for key in mode:
        #             shell=wb[key]  #表单名
        #             for i in range(2,shell.max_row+1):
        #                 row_data={}   #字典
        #                 row_data["case_id"] = shell.cell(i, 1).value
        #                 row_data["url"]=shell.cell(i,2).value
        #                 row_data["data"] = shell.cell(i, 3).value
        #                 row_data["title"] = shell.cell(i, 4).value
        #                 row_data["http_method"] = shell.cell(i, 5).value
        #                 row_data["expected"] = shell.cell(i, 6).value
        #                 test_data.append(row_data)
        #         return test_data

        # 通过配置文件经行过滤
        tel=getattr(GetData,"NoRegTel") #从GetData里面拿到了数据
        test_data=[] #列表
        for key in mode:
            shell=wb[key]  #表单名
            if mode[key]=="all":
                for i in range(2,shell.max_row+1):
                    row_data={}   #字典
                    row_data["case_id"] = shell.cell(i, 1).value
                    row_data["url"]=shell.cell(i,2).value
                    # 做手机号替换
                    if shell.cell(i,3).value.find("${tel_1}")!=-1:#有找到这个${tel_1}
                        row_data["data"]=shell.cell(i, 3).value.replace("${tel_1}",str(tel)) #找到后做替换
                    elif shell.cell(i,3).value.find("${tel}")!=-1:
                        row_data["data"] = shell.cell(i, 3).value.replace("${tel}", str(tel+1))
                    else:
                        row_data["data"] = shell.cell(i, 3).value
                    row_data["title"] = shell.cell(i, 4).value
                    row_data["http_method"] = shell.cell(i, 5).value
                    row_data["expected"] = shell.cell(i, 6).value
                    row_data["sheet_name"]=key
                    test_data.append(row_data)
                    cls.updata_tel(tel+2,file_name,"init")#更新手机号
            else:
                for id in mode[key]:
                    row_data = {}  #   字典
                    row_data["case_id"] = shell.cell(id + 1, 1).value
                    row_data["url"] = shell.cell(id + 1, 2).value
                    #做手机号替换
                    if shell.cell(id,3).value.find("${tel_1}")!=-1:#有找到这个${tel_1}
                        row_data["data"]=shell.cell(id, 3).value.replace("${tel_1}",str(tel)) #找到后做替换
                    elif shell.cell(id,3).value.find("${tel}")!=-1:
                        row_data["data"] = shell.cell(id, 3).value.replace("${tel}", str(tel+1))
                    else:
                        row_data["data"] = shell.cell(id, 3).value
                    row_data["title"] = shell.cell(id + 1, 4).value
                    row_data["http_method"] = shell.cell(id + 1, 5).value
                    row_data["expected"] = shell.cell(id + 1, 6).value
                    row_data["sheet_name"] = key
                    test_data.append(row_data)
                    cls.updata_tel(tel+2,file_name,"init")#更新手机号
        return test_data



    #往EXcel写入数据
    @staticmethod
    def write_back(file_name, sheet_name, i, result, TestResult): #创建写回数据
        wb=load_workbook(file_name) #参数化传参 传入文件路径
        sheet=wb[sheet_name]  #参数化传参 传入页签路径
        sheet.cell(i, 7).value=result
        sheet.cell(i, 8).value =TestResult
        wb.save(file_name)#保存结果

    @classmethod
    def updata_tel(cls,tel,filename,sheet_name):
        """更新excel里面指定位置的数据"""
        wb=load_workbook(filename)
        sheet=wb[sheet_name]
        sheet.cell(2,1).value=tel
        wb.save(filename)


if __name__ == '__main__':
    test_data=DoExcel().get_data(project_path.test_case_path)
    print(test_data)
    print(len(test_data))

