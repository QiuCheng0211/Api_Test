from openpyxl import load_workbook
from tools.read_config import ReadConfig  # 引入配置的py文件
from tools import project_path  # 引入公共路径的py文件
from tools.get_data import GetData  # 引入需要替换数据的py文件

""""
使用excel处理数据
输入：表格的路径 关联配置文件读取
"""


class DoExcel:

    @classmethod
    def get_data(cls, file_name):
        wb = load_workbook(file_name)  # 参数化传参 传入文件路径
        mode = eval(ReadConfig.get_config(project_path.case_config_path, "MODE", "mode"))  # 通过配置文件，读取表单名

        # 通过配置文件经行过滤
        tel = getattr(GetData, "NoRegTel")  # 从GetData里面拿到了需要做替换的备用数据
        test_data = []  # 定义一个空列表
        for key in mode:  # 遍历读取到的配置文件的表单名
            shell = wb[key]  # excel文件[表单名]
            if mode[key] == "all":  # 读取的是case.config配置文件的key的值  key等于"all"的时候则执行
                for i in range(2, shell.max_row + 1):  # 从第二行开始至获取最大行结束
                    row_data = {}  # 定义一个空字典
                    row_data["case_id"] = shell.cell(i, 1).value  # 第二行第一个值
                    row_data["url"] = shell.cell(i, 2).value  # 第二行第二个值
                    # 做手机号替换
                    if shell.cell(i, 3).value.find("${tel}") != -1:  # 有找到这个${tel_1}
                        row_data["data"] = shell.cell(i, 3).value.replace("${tel}", str(tel))  # 找到后做替换
                    elif shell.cell(i, 3).value.find("${admin_tel}") != -1:
                        row_data["data"] = shell.cell(i, 3).value.replace("${admin_tel}", str(getattr(GetData,"admin_tel")))
                    elif shell.cell(i,3).value.find("${loan_member_id}") != -1:
                        row_data["data"] = shell.cell(i, 3).value.replace("${loan_member_id}", str(getattr(GetData,"loan_member_id")))
                    elif shell.cell(i,3).value.find("${normal_tel}") != -1:
                        row_data["data"] = shell.cell(i, 3).value.replace("${normal_tel}", str(getattr(GetData,"normal_tel")))
                    elif shell.cell(i,3).value.find("${member_ID}") != -1:
                        row_data["data"] = shell.cell(i, 3).value.replace("${member_ID}", str(getattr(GetData,"member_ID")))
                    else:
                        row_data["data"] = shell.cell(i, 3).value
                    row_data["title"] = shell.cell(i, 4).value
                    row_data["http_method"] = shell.cell(i, 5).value
                    row_data["expected"] = shell.cell(i, 6).value
                    row_data["sheet_name"] = key
                    test_data.append(row_data) #将获取的每一行字典数据添加到列表里
                    cls.updata_tel(tel + 2, file_name, "init")  # 更新手机号
            else:
                for id in mode[key]:  # 读取的是case.config配置文件的key的值
                    row_data = {}  # 定义一个空字典
                    row_data["case_id"] = shell.cell(id + 1, 1).value# 第二行第一个值
                    row_data["url"] = shell.cell(id + 1, 2).value# 第二行第二个值
                    # 做手机号替换
                    if shell.cell(id + 1, 3).value.find("${tel_1}") != -1:  # 有找到这个${tel_1}
                        row_data["data"] = shell.cell(id + 1, 3).value.replace("${tel_1}", str(tel))  # 找到后做替换
                    elif shell.cell(id + 1, 3).value.find("${admin_tel}") != -1:
                        row_data["data"] = shell.cell(id + 1, 3).value.replace("${admin_tel}", str(getattr(GetData,"admin_tel")))
                    elif shell.cell(id + 1,3).value.find("${loan_member_id}")!=-1:
                        row_data["data"] = shell.cell(id + 1, 3).value.replace("${loan_member_id}", str(getattr(GetData,"loan_member_id")))
                    elif shell.cell(id + 1,3).value.find("${normal_tel}")!=-1:
                        row_data["data"] = shell.cell(id + 1, 3).value.replace("${normal_tel}", str(getattr(GetData,"normal_tel")))
                    elif shell.cell(id + 1,3).value.find("${member_ID}")!=-1:
                        row_data["data"] = shell.cell(id + 1, 3).value.replace("${member_ID}", str(getattr(GetData,"member_ID")))
                    else:
                        row_data["data"] = shell.cell(id, 3).value
                    row_data["title"] = shell.cell(id + 1, 4).value
                    row_data["http_method"] = shell.cell(id + 1, 5).value
                    row_data["expected"] = shell.cell(id + 1, 6).value
                    row_data["sheet_name"] = key
                    test_data.append(row_data)
                    cls.updata_tel(tel + 2, file_name, "init")  # 更新手机号
        return test_data

    # 往Excel写入数据
    @staticmethod
    def write_back(file_name, sheet_name, i, result, TestResult):  # 创建写回数据
        wb = load_workbook(file_name)  # 参数化传参 传入文件路径
        sheet = wb[sheet_name]  # 参数化传参 传入页签路径
        sheet.cell(i, 7).value = result  # 将传进来的result值进行回写Excel表格里
        sheet.cell(i, 8).value = TestResult # 将传进来的TestResult值进行回写Excel表格里
        wb.save(file_name)  # 保存结果

    @classmethod
    def updata_tel(cls, tel, filename, sheet_name):
        """更新excel里面指定位置的数据"""
        wb = load_workbook(filename)  # 参数化传参 传入文件路径
        sheet = wb[sheet_name]  # 参数化传参 传入页签路径
        sheet.cell(2, 1).value = tel  # 将传进来的tel值进行回写Excel表格里
        wb.save(filename)  # 保存结果


"""以下为测试本py文件"""
if __name__ == '__main__':
    test_data = DoExcel().get_data(project_path.test_case_path)
    print(test_data)
    print(len(test_data))
